#!/usr/bin/env python3
"""Regenerate jobs.xlsx (and jobs/README.md) from the md files and runs.yml.

Source of truth:
  jobs/<id>.md          — one file per active job, YAML frontmatter + free-form body
  jobs/closed/<id>.md   — one file per closed job, same schema plus date_closed /
                          closure_reason
  runs.yml              — one entry per scheduled run, drives the Log tab

jobs.xlsx and jobs/README.md are derived artifacts. They are fully rewritten on
every run; do not edit them by hand — changes will be clobbered.

Typical use:

    python regenerate_tracker.py

Exit codes:
  0  success
  1  validation error (missing frontmatter, bad schema, etc.)
"""
from __future__ import annotations

import datetime
import re
import sys
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
JOBS_DIR = ROOT / "jobs"
CLOSED_DIR = JOBS_DIR / "closed"
XLSX_PATH = ROOT / "jobs.xlsx"
RUNS_YML = ROOT / "runs.yml"
INDEX_MD = JOBS_DIR / "README.md"


# xlsx column schema — (header label, frontmatter key). Order matters.
ACTIVE_COLUMNS: list[tuple[str, str]] = [
    ("Job ID", "job_id"),
    ("Status", "status"),
    ("Priority", "priority"),
    ("Title", "title"),
    ("Employer", "employer"),
    ("Employer Type", "employer_type"),
    ("Category", "category"),
    ("Location", "location"),
    ("Loc Tier", "loc_tier"),
    ("Remote?", "remote"),
    ("Salary", "salary"),
    ("Date Posted", "date_posted"),
    ("Date Found", "date_found"),
    ("Last Checked", "last_checked"),
    ("Deadline", "deadline"),
    ("Apply URL", "apply_url"),
    ("Source", "source"),
    ("Research Flag", "research_flag"),
    ("Notes", "notes"),
    ("Details File", "details_file"),
]
CLOSED_COLUMNS: list[tuple[str, str]] = ACTIVE_COLUMNS + [
    ("Date Closed", "date_closed"),
    ("Reason", "closure_reason"),
]
LOG_COLUMNS: list[tuple[str, str]] = [
    ("Run Date", "run_date"),
    ("Run Type", "run_type"),
    ("New Jobs", "new_jobs"),
    ("Updated", "updated"),
    ("Closed", "closed"),
    ("Total Active", "total_active"),
    ("Queries Run", "queries_run"),
    ("Notes", "notes"),
]

# Fields that must be present in every job frontmatter.
REQUIRED_KEYS = {
    "job_id", "status", "priority", "title", "employer", "employer_type",
    "category", "location", "loc_tier", "remote", "date_found", "last_checked",
    "apply_url", "source",
}

FRONTMATTER_RE = re.compile(r"^---\s*\n(.*?)\n---\s*\n", re.DOTALL)


def parse_frontmatter(md_path: Path) -> dict:
    text = md_path.read_text(encoding="utf-8")
    m = FRONTMATTER_RE.match(text)
    if not m:
        raise ValueError(f"{md_path}: missing YAML frontmatter block")
    try:
        data = yaml.safe_load(m.group(1))
    except yaml.YAMLError as e:
        raise ValueError(f"{md_path}: invalid YAML frontmatter: {e}") from e
    if not isinstance(data, dict):
        raise ValueError(f"{md_path}: frontmatter must be a mapping")
    missing = REQUIRED_KEYS - set(data.keys())
    if missing:
        raise ValueError(f"{md_path}: missing required keys: {sorted(missing)}")
    return data


def load_jobs() -> tuple[list[dict], list[dict]]:
    active, closed = [], []

    for p in sorted(JOBS_DIR.glob("*.md")):
        if p.name.lower() == "readme.md":
            continue
        job = parse_frontmatter(p)
        job["details_file"] = f"jobs/{p.name}"
        if str(job.get("status", "")).lower() == "closed":
            raise ValueError(
                f"{p}: status=Closed but file is in jobs/, not jobs/closed/. Move it."
            )
        active.append(job)

    if CLOSED_DIR.exists():
        for p in sorted(CLOSED_DIR.glob("*.md")):
            job = parse_frontmatter(p)
            job["details_file"] = f"jobs/closed/{p.name}"
            if str(job.get("status", "")).lower() != "closed":
                raise ValueError(
                    f"{p}: file is in jobs/closed/ but status is not Closed"
                )
            closed.append(job)

    active.sort(key=lambda j: str(j["job_id"]))
    closed.sort(key=lambda j: str(j["job_id"]))
    return active, closed


def load_runs() -> list[dict]:
    if not RUNS_YML.exists():
        return []
    data = yaml.safe_load(RUNS_YML.read_text(encoding="utf-8")) or []
    if not isinstance(data, list):
        raise ValueError(f"{RUNS_YML}: top-level must be a list")
    return data


def coerce(value):
    """Convert a frontmatter value to something xlsx wants."""
    if value is None:
        return None
    if isinstance(value, bool):
        return "Y" if value else "N"
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.isoformat()[:10]
    return value


HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_FONT = Font(name="Arial", size=10, bold=True)
BODY_FONT = Font(name="Arial", size=10)
WRAP = Alignment(vertical="top", wrap_text=True)

COLUMN_WIDTHS = {
    "Job ID": 12, "Status": 10, "Priority": 9, "Title": 38, "Employer": 32,
    "Employer Type": 18, "Category": 16, "Location": 24, "Loc Tier": 8,
    "Remote?": 8, "Salary": 24, "Date Posted": 14, "Date Found": 12,
    "Last Checked": 12, "Deadline": 20, "Apply URL": 40, "Source": 20,
    "Research Flag": 8, "Notes": 60, "Details File": 24,
    "Date Closed": 12, "Reason": 50,
    "Run Date": 12, "Run Type": 12, "New Jobs": 10, "Updated": 10,
    "Closed": 10, "Total Active": 12, "Queries Run": 12,
}


def write_sheet(ws, columns, rows, freeze=True):
    for c, (label, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
    for r, row in enumerate(rows, start=2):
        for c, (_, key) in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=c, value=coerce(row.get(key)))
            cell.font = BODY_FONT
            cell.alignment = WRAP
    if freeze:
        ws.freeze_panes = "A2"
    for c, (label, _) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(c)].width = COLUMN_WIDTHS.get(label, 14)


def md_escape(value) -> str:
    if value is None:
        return ""
    return str(value).replace("|", "\\|").replace("\n", " ")


def write_index(active, closed):
    lines = [
        "# Jobs index",
        "",
        "_Auto-generated by `regenerate_tracker.py` — do not edit by hand._",
        "",
        "## Active",
        "",
        "| Job ID | Priority | Title | Employer | Location | Deadline |",
        "|---|---|---|---|---|---|",
    ]
    for j in active:
        lines.append(
            "| [{id}]({id}.md) | {pri} | {title} | {emp} | {loc} | {dl} |".format(
                id=md_escape(j["job_id"]),
                pri=md_escape(j.get("priority")),
                title=md_escape(j.get("title")),
                emp=md_escape(j.get("employer")),
                loc=md_escape(j.get("location")),
                dl=md_escape(j.get("deadline")),
            )
        )
    lines += [
        "",
        "## Closed",
        "",
        "| Job ID | Title | Employer | Closed | Reason |",
        "|---|---|---|---|---|",
    ]
    for j in closed:
        lines.append(
            "| [{id}](closed/{id}.md) | {title} | {emp} | {dc} | {reason} |".format(
                id=md_escape(j["job_id"]),
                title=md_escape(j.get("title")),
                emp=md_escape(j.get("employer")),
                dc=md_escape(j.get("date_closed")),
                reason=md_escape(j.get("closure_reason")),
            )
        )
    INDEX_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    try:
        active, closed = load_jobs()
        runs = load_runs()
    except ValueError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1

    wb = Workbook()
    ws_a = wb.active
    ws_a.title = "Active"
    write_sheet(ws_a, ACTIVE_COLUMNS, active)
    write_sheet(wb.create_sheet("Closed"), CLOSED_COLUMNS, closed)
    write_sheet(wb.create_sheet("Log"), LOG_COLUMNS, runs)
    wb.save(XLSX_PATH)

    write_index(active, closed)

    print(
        f"Wrote {XLSX_PATH.name}: {len(active)} active, {len(closed)} closed, "
        f"{len(runs)} runs."
    )
    print(f"Wrote {INDEX_MD.relative_to(ROOT)}.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
