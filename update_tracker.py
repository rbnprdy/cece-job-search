"""One-shot update of jobs.xlsx for the 2026-04-17 scheduled run."""
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font

WB_PATH = "jobs.xlsx"
RUN_DATE = "2026-04-17"

wb = load_workbook(WB_PATH)
active = wb["Active"]
closed = wb["Closed"]
log = wb["Log"]

active_headers = [c.value for c in active[1]]
closed_headers = [c.value for c in closed[1]]
log_headers = [c.value for c in log[1]]


def col_idx(headers, name):
    return headers.index(name) + 1


# IDs being moved to Closed with reason
to_close = {
    "260417-001": "Posting withdrawn (HEARCareers listing no longer active).",
    "260417-007": "Posting withdrawn (Mount Sinai req 3024126 returns 404; superseded by 260417-018 / req 3033391).",
    "260417-011": "Posting withdrawn (CHOP careers URL returns 410 Gone).",
    "260417-012": "Posting withdrawn (HEARCareers listing no longer active; direct email may still work).",
    "260417-016": "Application deadline (2026-02-27) passed.",
}

# Find rows to close on Active tab
id_col = col_idx(active_headers, "Job ID")
status_col = col_idx(active_headers, "Status")
last_checked_col = col_idx(active_headers, "Last Checked")
salary_col = col_idx(active_headers, "Salary")
deadline_col = col_idx(active_headers, "Deadline")

active_rows_by_id = {}
for r in range(2, active.max_row + 1):
    jid = active.cell(row=r, column=id_col).value
    if jid:
        active_rows_by_id[jid] = r


def copy_row_style(src_ws, src_row, dst_ws, dst_row, num_cols):
    """Copy cell values and formatting from one row to another row (possibly in different sheet)."""
    for c in range(1, num_cols + 1):
        src = src_ws.cell(row=src_row, column=c)
        dst = dst_ws.cell(row=dst_row, column=c)
        dst.value = src.value
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)


# 1) Move closed rows from Active → Closed tab
# Find next empty row in Closed sheet
closed_next = closed.max_row + 1
# skip trailing empties
while closed_next > 2 and closed.cell(row=closed_next - 1, column=1).value is None:
    closed_next -= 1

# For each to-close ID in deterministic order, copy the Active row data into Closed with status=Closed, Date Closed=today, Reason set
closed_date_col = col_idx(closed_headers, "Date Closed")
closed_reason_col = col_idx(closed_headers, "Reason")
closed_status_col = col_idx(closed_headers, "Status")
closed_last_checked_col = col_idx(closed_headers, "Last Checked")

# Build a template-style mapping from Active -> Closed based on header names
active_to_closed_map = {}
for h in active_headers:
    if h in closed_headers:
        active_to_closed_map[col_idx(active_headers, h)] = col_idx(closed_headers, h)

# Use row 2 of closed as style template
closed_template_row = 2

rows_to_delete_on_active = []
for jid in sorted(to_close.keys()):
    src_r = active_rows_by_id[jid]
    dst_r = closed_next
    # Copy formatting from template row of Closed
    for c in range(1, len(closed_headers) + 1):
        tmpl = closed.cell(row=closed_template_row, column=c)
        dst = closed.cell(row=dst_r, column=c)
        if tmpl.has_style:
            dst.font = copy(tmpl.font)
            dst.fill = copy(tmpl.fill)
            dst.border = copy(tmpl.border)
            dst.alignment = copy(tmpl.alignment)
            dst.number_format = tmpl.number_format
            dst.protection = copy(tmpl.protection)

    # Copy matching columns from Active
    for src_c, dst_c in active_to_closed_map.items():
        closed.cell(row=dst_r, column=dst_c).value = active.cell(row=src_r, column=src_c).value

    # Fill in Closed-specific fields
    closed.cell(row=dst_r, column=closed_status_col).value = "Closed"
    closed.cell(row=dst_r, column=closed_last_checked_col).value = RUN_DATE
    closed.cell(row=dst_r, column=closed_date_col).value = RUN_DATE
    closed.cell(row=dst_r, column=closed_reason_col).value = to_close[jid]

    rows_to_delete_on_active.append(src_r)
    closed_next += 1

# Delete Active rows in reverse order so row indexes don't shift
for r in sorted(rows_to_delete_on_active, reverse=True):
    active.delete_rows(r, 1)

# Re-index
active_rows_by_id = {}
for r in range(2, active.max_row + 1):
    jid = active.cell(row=r, column=id_col).value
    if jid:
        active_rows_by_id[jid] = r

# 2) Update remaining Active rows
# Update Last Checked for all remaining
for jid, r in active_rows_by_id.items():
    active.cell(row=r, column=last_checked_col).value = RUN_DATE

# Specific updates
if "260417-002" in active_rows_by_id:
    r = active_rows_by_id["260417-002"]
    active.cell(row=r, column=salary_col).value = "$83,241 – $133,192"
    active.cell(row=r, column=deadline_col).value = "2026-05-23"

if "260417-009" in active_rows_by_id:
    r = active_rows_by_id["260417-009"]
    active.cell(row=r, column=salary_col).value = "$69,525 – $131,654 (currently posted internal-only)"

# 3) Add 4 new rows: 260417-017, 018, 019, 020
new_rows = [
    {
        "Job ID": "260417-017",
        "Status": "New",
        "Priority": "Medium",
        "Title": "Audiologist, Clinical Specialist – Otolaryngology",
        "Employer": "Mount Sinai Health System",
        "Employer Type": "Academic",
        "Category": "Clinical-Admin",
        "Location": "New York, NY",
        "Loc Tier": "1",
        "Remote?": "No",
        "Salary": "Not listed",
        "Date Posted": "2026 (active)",
        "Date Found": RUN_DATE,
        "Last Checked": RUN_DATE,
        "Deadline": None,
        "Apply URL": "https://careers.mountsinai.org/jobs/3033433?lang=en-us",
        "Source": "Mount Sinai Careers",
        "Research Flag": "Y",
        "Notes": "Tier-1 NYC AMC (Icahn School of Medicine). 'Clinical Specialist' within Otolaryngology-HNS — typically senior-scope/subspecialty, teaching, cross-team consult. Adjacent to Mount Sinai's Center for Hearing & Balance research programs. PSLF-eligible. Adjacent-role flag for Cece.",
        "Details File": "jobs/260417-017.md",
    },
    {
        "Job ID": "260417-018",
        "Status": "New",
        "Priority": "Low",
        "Title": "Doctor of Audiology (NYEE, Full-time Days)",
        "Employer": "Mount Sinai Health System",
        "Employer Type": "Academic",
        "Category": "Clinical",
        "Location": "New York, NY",
        "Loc Tier": "1",
        "Remote?": "No",
        "Salary": "Not listed",
        "Date Posted": "2026 (active)",
        "Date Found": RUN_DATE,
        "Last Checked": RUN_DATE,
        "Deadline": None,
        "Apply URL": "https://careers.mountsinai.org/jobs/3033391?lang=en-us",
        "Source": "Mount Sinai Careers",
        "Research Flag": "N",
        "Notes": "Tier-1 NYC AMC. Doctor of Audiology at New York Eye and Ear Infirmary — full-time days. Appears to be the replacement req for the closed 260417-007 (Mount Sinai 3024126). PSLF-eligible. Pure clinical, surfaced per Tier-1 rule.",
        "Details File": "jobs/260417-018.md",
    },
    {
        "Job ID": "260417-019",
        "Status": "New",
        "Priority": "Low",
        "Title": "Audiologist – Full Time, Days",
        "Employer": "NewYork-Presbyterian (Brooklyn Methodist Hospital)",
        "Employer Type": "Hospital-NonProfit",
        "Category": "Clinical",
        "Location": "Brooklyn, NY",
        "Loc Tier": "1",
        "Remote?": "No",
        "Salary": "Not listed (NYP competitive; verify at apply page)",
        "Date Posted": "2026-01-23",
        "Date Found": RUN_DATE,
        "Last Checked": RUN_DATE,
        "Deadline": None,
        "Apply URL": "https://careers.nyp.org/job/brooklyn/audiologist-full-time-days/19715/90947179360",
        "Source": "NYP Careers",
        "Research Flag": "N",
        "Notes": "Tier-1 NYC non-profit (NYP Brooklyn Methodist is part of NewYork-Presbyterian; teaching hospital for Columbia Vagelos & Weill Cornell). Posted 2026-01-23 so likely still accepting applications. PSLF-eligible. Pure clinical (diagnostic + rehab, adult & pediatric).",
        "Details File": "jobs/260417-019.md",
    },
    {
        "Job ID": "260417-020",
        "Status": "New",
        "Priority": "Low",
        "Title": "Audiologist",
        "Employer": "Kaiser Permanente (South San Francisco Medical Offices)",
        "Employer Type": "Hospital-NonProfit",
        "Category": "Clinical",
        "Location": "South San Francisco, CA",
        "Loc Tier": "1",
        "Remote?": "No",
        "Salary": "$120,400 – $155,760 / yr",
        "Date Posted": "2026 (active)",
        "Date Found": RUN_DATE,
        "Last Checked": RUN_DATE,
        "Deadline": "Rolling",
        "Apply URL": "https://www.kaiserpermanentejobs.org/job/south-san-francisco/audiologist/641/80366214896",
        "Source": "Kaiser Permanente Jobs",
        "Research Flag": "N",
        "Notes": "Tier-1 SF Bay. Kaiser NorCal integrated non-profit — typically PSLF-eligible (verify specific Kaiser entity at apply). Full-time day shift (M–F, 8:30a–5:30p); Head & Neck Audiology dept. Min 1 yr post-licensure — within Cece's range. Pure clinical but high salary band and predictable schedule.",
        "Details File": "jobs/260417-020.md",
    },
]

active_template_row = 2  # row to copy style from
active_next = active.max_row + 1
# skip trailing empties
while active_next > 2 and active.cell(row=active_next - 1, column=1).value is None:
    active_next -= 1

for new in new_rows:
    for c in range(1, len(active_headers) + 1):
        tmpl = active.cell(row=active_template_row, column=c)
        dst = active.cell(row=active_next, column=c)
        if tmpl.has_style:
            dst.font = copy(tmpl.font)
            dst.fill = copy(tmpl.fill)
            dst.border = copy(tmpl.border)
            dst.alignment = copy(tmpl.alignment)
            dst.number_format = tmpl.number_format
            dst.protection = copy(tmpl.protection)
    for h, v in new.items():
        active.cell(row=active_next, column=col_idx(active_headers, h)).value = v
    active_next += 1

# 4) Append Log row
log_next = log.max_row + 1
while log_next > 2 and log.cell(row=log_next - 1, column=1).value is None:
    log_next -= 1

log_template_row = 2
log_values = {
    "Run Date": RUN_DATE,
    "Run Type": "Scheduled",
    "New Jobs": 4,
    "Updated": 7,
    "Closed": 5,
    "Total Active": 11,
    "Queries Run": 30,
    "Notes": "Second scheduled run. 4 new (Mount Sinai Clinical Specialist, Mount Sinai NYEE DofA replacement, NYP Brooklyn Methodist, Kaiser S. SF). Closed 5 (UCSF HEARCareers withdrawn; Mount Sinai 3024126 replaced; CHOP 410; Hopkins HEARCareers withdrawn; VA San Diego deadline passed). Salary refreshes on BCH ($83.2k–$133.2k, deadline 5/23) and SFUSD ($69.5k–$131.7k, internal-only).",
}
for c in range(1, len(log_headers) + 1):
    tmpl = log.cell(row=log_template_row, column=c)
    dst = log.cell(row=log_next, column=c)
    if tmpl.has_style:
        dst.font = copy(tmpl.font)
        dst.fill = copy(tmpl.fill)
        dst.border = copy(tmpl.border)
        dst.alignment = copy(tmpl.alignment)
        dst.number_format = tmpl.number_format
        dst.protection = copy(tmpl.protection)
for h, v in log_values.items():
    log.cell(row=log_next, column=col_idx(log_headers, h)).value = v

wb.save(WB_PATH)
print("Saved.")
print(f"Active rows now: {active.max_row}")
print(f"Closed rows now: {closed.max_row}")
print(f"Log rows now: {log.max_row}")
