"""Update jobs.xlsx for 2026-04-17 second run."""
from openpyxl import load_workbook
from openpyxl.styles import Font
from copy import copy

WORKSPACE = "/sessions/amazing-relaxed-sagan/mnt/cece-job-search"
XLSX = f"{WORKSPACE}/jobs.xlsx"

wb = load_workbook(XLSX)
active = wb["Active"]
log = wb["Log"]

# ---- Update existing 4 rows (refresh Last Checked, update salary/notes where applicable) ----
# Columns: 1=JobID ... 11=Salary, 14=LastChecked, 19=Notes
# Row 2: 260417-001 UCSF Sr Audiologist
# Row 3: 260417-002 Boston Children's Audiologist II
# Row 4: 260417-007 Mount Sinai NYEE
# Row 5: 260417-008 Scripps

today = "2026-04-17"

# Row 2 — UCSF: just update Last Checked
active.cell(row=2, column=14, value=today)

# Row 3 — Boston Children's: update Salary + Last Checked + Notes
active.cell(row=3, column=11, value="$76,600 – $107,000")
active.cell(row=3, column=14, value=today)
active.cell(row=3, column=19, value=(
    "Tier-1 non-profit, Harvard-affiliated. 40+ audiologist team, pediatric specialty "
    "with strong research arm (Dr. Amanda Griffin directs audiology research, focus on "
    "cochlear implants). Includes audiological testing of research patients in gene-therapy "
    "trials. Stated req: min 4 yrs pediatric audiology experience — borderline for Cece's "
    "2–3 yrs post-AuD (AuD training years may count; worth inquiring)."
))

# Row 4 — Mount Sinai NYEE: Last Checked
active.cell(row=4, column=14, value=today)

# Row 5 — Scripps: update Salary + Last Checked + Notes
active.cell(row=5, column=11, value="$54.24 – $78.66 / hr")
active.cell(row=5, column=14, value=today)
active.cell(row=5, column=19, value=(
    "Tier-1 San Diego non-profit. Clinical only. Min 1 yr licensed experience — within "
    "Cece's range. Hourly $54.24–$78.66 per aggregator. Surfaced per 'always include "
    "Tier-1 clinical' rule."
))

# ---- New rows (append at rows 6..13) ----
# Schema order: JobID, Status, Priority, Title, Employer, EmployerType, Category, Location,
#               LocTier, Remote?, Salary, DatePosted, DateFound, LastChecked, Deadline,
#               ApplyURL, Source, ResearchFlag, Notes, DetailsFile
new_rows = [
    (
        "260417-009", "New", "Medium",
        "Audiologist (2025-2026)", "San Francisco Unified School District",
        "Public-School", "Clinical", "San Francisco, CA", "1", "No",
        "Not listed (SFUSD certificated scale)", "2026 (active)", today, today,
        "Check posting",
        "https://www.edjoin.org/Home/JobPosting/2053232",
        "EDJOIN (SFUSD)", "N",
        "Matches top priority: public-school audiology, Tier-1 SF, PSLF-eligible. Educational "
        "audiologist for ~50,000-student district; FM/HAT systems, IEP/CSE collaboration. "
        "Reframes clinical work around education.",
        "jobs/260417-009.md",
    ),
    (
        "260417-010", "New", "Medium",
        "Audiologist", "NYC Department of Education (NYC Public Schools)",
        "Public-School", "Clinical", "New York, NY", "1", "No",
        "UFT certificated scale (not listed)", "Recurring", today, today,
        "Rolling / per-session",
        "https://www.schools.nyc.gov/careers/other-jobs-in-schools/audiologists",
        "NYC DOE", "N",
        "Matches top priority: public-school audiology, Tier-1 NYC, UFT + PSLF-eligible. "
        "District 75 + five-borough CSE work. Distinct from VA clinical; worth surfacing as "
        "baseline NYC anchor.",
        "jobs/260417-010.md",
    ),
    (
        "260417-011", "New", "Medium",
        "Pediatric Audiologist", "Children's Hospital of Philadelphia (CHOP)",
        "Hospital-NonProfit", "Clinical-Research", "Philadelphia, PA", "2", "No",
        "~$70k–$87k typical (Indeed)", "2026 (active)", today, today,
        "Not posted",
        "https://careers.chop.edu/us/en/job/1019766/Pediatric-Audiologist",
        "CHOP Careers", "Y",
        "Tier-2 Philly but CHOP role explicitly lists teaching + research as additional "
        "responsibilities — rare hybrid shape at an AMC. ~45-audiologist dept, strong "
        "pediatric hearing-loss research via CHOP Research Institute; Penn-affiliated. "
        "PSLF-eligible.",
        "jobs/260417-011.md",
    ),
    (
        "260417-012", "New", "Medium",
        "Clinical Audiologist (adult & pediatric)", "Johns Hopkins",
        "Academic", "Clinical", "Baltimore, MD", "2", "No",
        "Not listed", "2026 (active)", today, today,
        "Not posted (start Jan 2026)",
        "https://hearcareers.audiology.org/jobs/21597843/johns-hopkins-clinical-audiologist",
        "AAA HEARCareers", "Y",
        "Tier-2 DC metro (Baltimore). Top-tier AMC with research-active Div. of Audiology "
        "in Dept of Otolaryngology-HNS. Two openings (1 adult, 1 pediatric). Direct-email "
        "application to cryan1@jhmi.edu. PSLF-eligible. Adjacent role: ample research "
        "collaboration potential.",
        "jobs/260417-012.md",
    ),
    (
        "260417-013", "New", "Medium",
        "Full-time Audiologist (MEE)", "Mass Eye and Ear (Mass General Brigham)",
        "Academic", "Clinical-Research", "Boston, MA", "1", "No",
        "~$71k–$101k (MGB Workday)", "2026 (active)", today, today,
        "Not posted",
        "https://massgeneralbrigham.wd1.myworkdayjobs.com/en-US/MGBExternal/job/Full-time-Audiologist---MEE_RQ4026190",
        "MGB Careers (Workday)", "Y",
        "Tier-1 Boston, Harvard-affiliated (Harvard Otolaryngology based at MEE). Research "
        "ops team of 20+ clinical coordinators. Dr. Julie Arenberg directs Audiology "
        "Research & Education and co-launched MGH IHP AuD program — teaching opportunities "
        "possible. PSLF-eligible.",
        "jobs/260417-013.md",
    ),
    (
        "260417-014", "New", "Low",
        "Audiologist", "Rady Children's Hospital-San Diego",
        "Hospital-NonProfit", "Clinical", "San Diego, CA", "1", "No",
        "~$48–$66/hr (per-diem listing)", "2026 (active)", today, today,
        "Rolling (FT/casual/per-diem)",
        "https://jobs.rchsd.org/job/san-diego/audiologist-audiology/1717/93187320528",
        "Rady Children's Careers", "N",
        "Tier-1 San Diego pediatric non-profit, PSLF-eligible. Multiple openings (FT, casual, "
        "per-diem). Pure clinical but surfaced per Tier-1 rule.",
        "jobs/260417-014.md",
    ),
    (
        "260417-015", "New", "Low",
        "Audiologist I", "Stanford Health Care (Ear Institute)",
        "Academic", "Clinical", "Palo Alto, CA", "1", "No",
        "Not listed", "2026 (active)", today, today,
        "Not posted",
        "https://careers.stanfordhealthcare.org/",
        "Stanford Health Care", "N",
        "Tier-1 Bay Area AMC (Stanford OHNS). Entry/early-career band — reasonable fit for "
        "2–3 yrs post-AuD. Adjacent to Stanford Initiative to Cure Hearing Loss → research "
        "access via faculty collab possible. PSLF-eligible.",
        "jobs/260417-015.md",
    ),
    (
        "260417-016", "New", "Low",
        "Audiologist", "VA San Diego Healthcare System",
        "VA-Federal", "Clinical", "San Diego, CA (La Jolla)", "1", "No",
        "GS scale (per USAJobs)", "2026 (active)", today, today,
        "2026-02-27 (past — verify next run)",
        "https://www.usajobs.gov/job/857821800",
        "USAJobs", "N",
        "Tier-1 San Diego VA, PSLF-eligible, familiar environment (VA). CAVEAT: listed "
        "deadline 02/27/2026 already passed — listing may be closed or close soon. Watch "
        "for repost next run.",
        "jobs/260417-016.md",
    ),
]

# Append starting at row 6 (Active currently has rows 2-5 populated with 4 jobs)
start_row = 6
# Get style reference from existing data row (row 2)
for i, rowdata in enumerate(new_rows):
    r = start_row + i
    for c, val in enumerate(rowdata, start=1):
        cell = active.cell(row=r, column=c, value=val)
        # Match body font (Arial 10)
        cell.font = Font(name="Arial", size=10)

# ---- Append Log row ----
log_row = (
    "2026-04-17", "Scheduled", 8, 4, 0, 12, 21,
    "Second run (same day as bootstrap). Added 8 new jobs (2 school-audiology, 2 AMC "
    "hybrid clinical+research, 1 AMC adjacent, 3 Tier-1 clinical). Refreshed all 4 "
    "existing Active jobs (all still live; updated salary for BCH and Scripps). "
    "WebFetch blocked for most target domains — relied on WebSearch. Running today's "
    "date treated as 2026-04-17 per env."
)
log_start = log.max_row + 1
for c, val in enumerate(log_row, start=1):
    cell = log.cell(row=log_start, column=c, value=val)
    cell.font = Font(name="Arial", size=10)

wb.save(XLSX)
print(f"Saved. Active rows: {active.max_row}, Closed rows: {wb['Closed'].max_row}, Log rows: {log.max_row}")
